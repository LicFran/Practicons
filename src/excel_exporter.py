#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Módulo exportador de Excel para guardar datos extraídos de presupuestos de construcción
"""

import os
import logging
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from src.config import DEFAULT_SHEET_NAME, TABLE_HEADERS

logger = logging.getLogger(__name__)


class ExcelExporter:
    """
    Clase para exportar datos extraídos de presupuestos de construcción a Excel
    """
    
    def __init__(self, data, output_path):
        """
        Inicializa el exportador de Excel
        
        Args:
            data (dict): Datos extraídos del PDF con metadatos, tablas, etc.
            output_path (str or Path): Ruta al archivo Excel de salida
        """
        self.data = data
        self.output_path = Path(output_path)
        self.workbook = None
        
        # Definir estilos para el Excel
        self.header_font = Font(bold=True, size=12, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Bordes para las celdas
        self.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
    
    def export(self):
        """
        Exporta los datos extraídos a un archivo Excel formateado
        
        Returns:
            Path: Ruta al archivo Excel generado
        """
        logger.info(f"Exportando datos a Excel: {self.output_path}")
        
        # Crear un nuevo libro de trabajo
        self.workbook = openpyxl.Workbook()
        
        # Eliminar la hoja predeterminada
        default_sheet = self.workbook.active
        self.workbook.remove(default_sheet)
        
        # Crear hojas para diferentes tipos de datos
        self._create_proyecto_sheet()     # Hoja para datos del proyecto
        self._create_materiales_sheet()   # Hoja para listado de materiales
        
        # Guardar el archivo Excel
        self.workbook.save(self.output_path)
        logger.info(f"Archivo Excel guardado con éxito: {self.output_path}")
        
        return self.output_path
    
    def _create_proyecto_sheet(self):
        """Crear hoja para datos del proyecto"""
        # Obtener metadatos y datos del proyecto
        metadata = self.data.get("metadata", {})
        datos_proyecto = self.data.get("datos_proyecto", {})
        
        # Registrar los datos recibidos
        logger.info(f"Datos para hoja Proyecto - Metadata: {metadata}")
        logger.info(f"Datos para hoja Proyecto - Datos proyecto: {datos_proyecto}")
        
        # Combinar datos regulares y mejorados por IA
        combined_data = {**metadata}
        if datos_proyecto:
            combined_data.update({k: v for k, v in datos_proyecto.items() if v and not combined_data.get(k)})
        
        logger.info(f"Datos combinados para hoja Proyecto: {combined_data}")
        
        # Crear hoja de proyecto
        sheet = self.workbook.create_sheet("Proyecto")
        
        # Definir los encabezados según lo solicitado
        headers = [
            "Nombre Proyecto", 
            "Cliente", 
            "Celular", 
            "Tel Fijo", 
            "Direccion", 
            "Email", 
            "Fecha",
            "Orden Trabajo",
            "Total Materiales",
            "Total Mano de Obra",
            "Total Proyecto"
        ]
        
        # Escribir encabezados
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
            # Ajustar ancho de columna
            sheet.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 2, 15)
        
        # Verificar si hay datos
        if not combined_data:
            sheet.cell(row=2, column=1, value="No se encontraron datos del proyecto")
            return
        
        # Mapeo de nombres de campo a los encabezados
        field_map = {
            "nombre_proyecto": 0,  # Nombre Proyecto
            "project_name": 0,     # Alternativa
            "cliente": 1,          # Cliente
            "client": 1,           # Alternativa
            "celular": 2,          # Celular
            "phone": 2,            # Alternativa
            "telefono_fijo": 3,    # Tel Fijo
            "phone_fixed": 3,      # Alternativa
            "direccion": 4,        # Direccion
            "address": 4,          # Alternativa
            "email": 5,            # Email
            "e-mail": 5,           # Alternativa
            "fecha": 6,            # Fecha
            "date": 6,             # Alternativa
            "orden_trabajo": 7,    # Orden Trabajo
            "work_order": 7,       # Alternativa
            "total_materiales": 8, # Total Materiales
            "total_material": 8,   # Alternativa
            "total_mano_obra": 9,  # Total Mano de Obra
            "total_mano_de_obra": 9, # Alternativa
            "total_proyecto": 10,  # Total Proyecto
            "total_general": 10,   # Alternativa
            "total_amount": 10     # Alternativa
        }
        
        # Preparar datos para la fila
        row_data = [""] * len(headers)
        
        # Rellenar con datos disponibles
        for key, value in combined_data.items():
            key_lower = key.lower()
            if key_lower in field_map and value:
                row_data[field_map[key_lower]] = value
        
        # Escribir datos en la fila 2
        for col_idx, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=2, column=col_idx, value=value)
            cell.border = self.border
            
            # Formatear valores numéricos en las columnas de totales
            if col_idx in [9, 10, 11]:  # Columnas de totales
                try:
                    if isinstance(value, str):
                        # Limpiar y convertir a número si es posible
                        clean_value = value.replace(",", "").replace("$", "").strip()
                        if clean_value:
                            cell.value = float(clean_value)
                            cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")
                except ValueError:
                    # Si no se puede convertir a número, dejarlo como está
                    pass
        
        # Aplicar autofilter
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}2"
        
        # Congelar la fila de encabezados
        sheet.freeze_panes = "A2"
    
    def _create_materiales_sheet(self):
        """Crear hoja para listado y precios de materiales"""
        # Obtener lista de materiales
        materiales = self.data.get("materiales", [])
        logger.info(f"Datos de materiales extraídos: {materiales}")
        
        if not materiales:
            # Intentar obtener datos de tablas si no hay materiales específicos
            tables = self.data.get("tables", [])
            all_data = []
            for table in tables:
                all_data.extend(table.get("data", []))
            
            logger.info(f"Datos de tablas para materiales: {all_data}")
            
            if not all_data:
                # Crear una hoja vacía si no hay datos
                sheet = self.workbook.create_sheet("Materiales")
                sheet["A1"] = "No se encontraron datos de materiales"
                return
            
            # Usar datos de tablas como materiales
            materiales = all_data
        
        # Crear hoja de materiales
        sheet = self.workbook.create_sheet("Materiales")
        
        # Definir encabezados para la hoja de materiales
        headers = ["Material", "Unidades", "Precio Unitario", "Precio Total"]
        
        # Escribir encabezados
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
            # Ajustar ancho de columna
            sheet.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 2, 15)
        
        # Escribir datos de materiales
        for row_idx, item in enumerate(materiales, start=2):
            if isinstance(item, dict):
                # Asignar valores a las columnas correspondientes
                values = [
                    item.get("material", item.get("nombre", "")),
                    item.get("units", item.get("unidades", "")),
                    item.get("unit_price", item.get("precio_unitario", "")),
                    item.get("total_price", item.get("precio_total", ""))
                ]
                
                for col_idx, value in enumerate(values, start=1):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = self.border
                    
                    # Formatear valores numéricos
                    if col_idx in [3, 4]:  # Columnas de precios
                        try:
                            if isinstance(value, str):
                                # Limpiar y convertir a número si es posible
                                clean_value = value.replace(",", "").replace("$", "").strip()
                                if clean_value:
                                    cell.value = float(clean_value)
                                    cell.number_format = "#,##0.00"
                            cell.alignment = Alignment(horizontal="right")
                        except ValueError:
                            # Si no se puede convertir a número, dejarlo como está
                            pass
            elif isinstance(item, list):
                # Si es una lista, interpretar las primeras columnas como materiales
                for col_idx, value in enumerate(item[:4], start=1):  # Limitar a las primeras 4 columnas
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = self.border
                    if col_idx in [3, 4]:
                        cell.alignment = Alignment(horizontal="right")
        
        # Ajustar el ancho de las columnas
        for col_idx in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col_idx)
            for row_idx in range(1, len(materiales) + 2):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = max(max_length + 2, 15)
            sheet.column_dimensions[column].width = min(adjusted_width, 50)  # Limitar el ancho máximo
        
        # Aplicar autofilter y congelar encabezados
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(materiales) + 1}"
        sheet.freeze_panes = "A2"
    